import os
from datetime import date, timedelta, datetime
from marketorestpython.client import MarketoClient
from marketorestpython.batch import MarketoClientBatch
from simple_salesforce import Salesforce, SFType, SalesforceResourceNotFound
import json

# Load credentials from json file
with open("marketo_batch_settings.json", "r") as file:  
    settings = json.load(file)

server = settings['deploy']
mkto_config = settings['config']

def get_sf_config(config_path):
    # the secrets file contains the following:
    # username
    # password
    # token
    if not os.path.exists(config_path):
        sf_cnf_data = {
            'username':'<<>>',
            'passwd': '<<>>',
            'token': '<<>>',
            'uri': '<<>>'
        }
        with open(config_path, 'w') as outfile:
            json.dump(sf_cnf_data, outfile)

        print('Created empty config file [{}], please update and run again!'.format(config_path))

    sf_cnf_data = json.load(open(config_path))

    if sf_cnf_data['passwd'] == '<<>>':
        import getpass
        sf_cnf_data['passwd'] = getpass.getpass(" Please enter you password:")
    print('Config loaded... {}'.format(datetime.now()))
    return sf_cnf_data

def connect_sf(config_data):
    sf = Salesforce(username=config_data['username'], password=config_data['passwd'], security_token=config_data['token'], sandbox=True)
    return sf

def get_mkto_fields():
    server = 'Prod'

    munchkin_id = mkto_config[server]['munchkin_id']         ### Enter Munchkin ID
    client_id = mkto_config[server]['client_id']             ### enter client ID (find in Admin > LaunchPoint > view details)
    client_secret = mkto_config[server]['client_secret']     ### enter client secret (find in Admin > LaunchPoint > view details)

    mc = MarketoClient(munchkin_id, client_id, client_secret)
    resp = mc.describe()

    field_md = []
    for r in resp:
        field = {}
        field['id'] = r['id']
        field['name'] = r['displayName']
        field['api'] = r['rest']['name']
        #field['soap'] = r['soap']['name']
        field['type'] = r['dataType']
        if 'length' in r:
            field['length'] = r['length']
        field_md.append(field)
    return field_md

def process_sf_fields(fields, label_field, name_field, types=None):
    field_md = []
    for field in fields:
        md_ = {}
        md_['label'] = field[label_field]
        md_['name'] = field[name_field]
        for attribute in field:
            if types is None or attribute in types:
                md_[attribute] = field[attribute]
        field_md.append(md_)

    return field_md

def get_sf_field(sf, obj_dict):
    field_md = []
    for key in obj_dict:
        print('\t{}'.format(key))
        sf_obj = obj_dict[key]
        obj_name = sf_obj['objectName']
        
        # get
        try:
            #obj_metadata = SFType(obj_name, sf.session_id, sf.sf_instance, sf.sf_version, sf.proxies).metadata()
            obj_describe = SFType(obj_name, sf.session_id, sf.sf_instance, sf.sf_version, sf.proxies).describe()
        except SalesforceResourceNotFound:
            obj_error = 'SalesforceResourceNotFound'
            continue
    
        # extract the fields:
        fields = obj_describe['fields']
        types = ['deprecatedAndHidden', 'picklistValues', 'soapType', 'type', 'length', 'defaultValue', 'defaultValueFormula', 'externalId', 'calculated', 'calculatedFormula', 'compoundFieldName', 'controllerName']
        resp = process_sf_fields(fields, 'name', 'label', types)
        print('\tFound {} fields.'.format(len(resp)))
        for r in resp:
            obj_md = {}
            obj_md['systemName'] = 'salesforce'
            obj_md['tablename'] = obj_describe['name']
            obj_md['tablelabel'] = obj_describe['label']
            obj_md['tablednh'] = obj_describe['deprecatedAndHidden']

            obj_md['fieldname'] =  r['name']
            obj_md['fieldlabel'] = r['label']
            obj_md['fielddnh'] = r['deprecatedAndHidden']

            obj_md['datatype'] = r['type']

            print('\t{}.{} => {} [F:{}|C:{}|A:{}]'.format(obj_md['tablename'], obj_md['fieldname'], len(r['picklistValues']), r['calculatedFormula'], r['compoundFieldName'], r['controllerName']))
            #print(r)

            buiness_rule = ''
            if len(r['picklistValues']) > 0:
                buiness_rule = 'PickList: '
                for key in r['picklistValues']:
                    buiness_rule += '{}|'.format(key['label'])
            obj_md['businessrule'] = buiness_rule

            validation_rule = ''
            if r['calculatedFormula'] is not None:
                validation_rule += r['calculatedFormula']

            if r['controllerName'] is not None:
                validation_rule += r['controllerName']

            if r['compoundFieldName'] is not None:
                validation_rule += r['compoundFieldName']

            obj_md['validationRule'] = validation_rule

            field_md.append(obj_md)

    return field_md

sf_dict = {
        'Account': {'name': '', 'objectName': 'Account'},
        'Booking': {'name': '', 'objectName': 'Booking__c'},
        'Booking List': {'name': '', 'objectName': 'Passenger_Booking_Detail__c'},
        'Campaign': {'name': '', 'objectName': 'Campaign'},
        'Contact': {'name': '', 'objectName': 'Contact'},
        'Contract': {'name': '', 'objectName': 'Contract'},
        'Invoices': {'name': '', 'objectName': 'Invoices__c'},
        'Lead': {'name': '', 'objectName': 'Lead'},
        'Opportunity': {'name': '', 'objectName': 'Opportunity'},
        'Task': {'name': '', 'objectName': 'Task'},
        'Event': {'name': '', 'objectName': 'Event'}
    }

field_md_mkto = get_mkto_fields()
print('found {} field in marketo'.format(len(field_md_mkto)))

sf_fields = []
sf_config_path = r'sf.secrets.json'
sf_config_data = get_sf_config(sf_config_path)
print('loaded sf config...')

sf = connect_sf(sf_config_data)
print('connected to sf...')

print('grabbing metadata...')
sf_fields.extend(get_sf_field(sf, sf_dict))
print('\tFound {} fields'.format(len(sf_fields)))

import openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

import string
wb_file_path = 'metadata.xlsx'

print('creating workbook...')
wb = openpyxl.Workbook()

columns = ['Name', 'System Name', 'Table Name', 'Field Name', 'Datatype', 'Business Values', 'Validation Rule', 'Created by', 'Created date', 'Definition', 'Purpose', 'Associated Business Process']

print('saving marketo...')
ws = wb.create_sheet('Marketo')
row = 2
col = 1
for c in columns:
    ws.cell(row = row, column=col).value=c

for r in field_md_mkto:
    row += 1
    ws.cell(row = row, column=1).value=r['name']
    ws.cell(row = row, column=2).value='marketo'
    ws.cell(row = row, column=3).value='leads'
    ws.cell(row = row, column=4).value=r['api']
    ws.cell(row = row, column=5).value=r['type']

print('saving saleforce...')
ws = wb.create_sheet('Salesforce')
row = 2
col = 1
for c in columns:
    ws.cell(row = row, column=col).value=c

for r in sf_fields:
    row += 1
    ws.cell(row = row, column=1).value=r['fieldlabel']
    ws.cell(row = row, column=2).value=r['systemName']
    ws.cell(row = row, column=3).value=r['tablelabel']
    ws.cell(row = row, column=4).value=r['fieldname']
    ws.cell(row = row, column=5).value=r['datatype']
    ws.cell(row = row, column=6).value=r['businessrule']
    ws.cell(row = row, column=7).value=r['validationRule']
    #ws.cell(row = row, column=8).value=r['name']
    #ws.cell(row = row, column=9).value=r['name']
    #ws.cell(row = row, column=10).value=r['name']
    #ws.cell(row = row, column=11).value=r['name']
    #ws.cell(row = row, column=12).value=r['name']

print('saving...')    
wb.save(wb_file_path)
print('completed!')

